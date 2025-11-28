from flask import Blueprint, send_file, request
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import io

from . import export_bp

@export_bp.route("/exportar", methods=["POST"])
def exportar_excel():
    cantidades = request.form.getlist("qty_values[]")
    nombres = request.form.getlist("names[]")
    ubicaciones = request.form.getlist("locations[]")
    categorias = request.form.getlist("categories[]")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    header_fill = PatternFill(start_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    ws.append(["Producto", "Cantidad", "Ubicación", "Categoría"])

    # Estilo encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Rellenar filas
    for n, c, u, cat in zip(nombres, cantidades, ubicaciones, categorias):
        try:
            c_int = int(float(c))
        except:
            c_int = 0
        
        # Color según cantidad
        if c_int == 0:
            fill = PatternFill(start_color="D3D3D3", fill_type="solid") # negro
        elif c_int < 3:
            fill = PatternFill(start_color="FF9999", fill_type="solid") # rojo
        elif c_int < 7:
            fill = PatternFill(start_color="FFF59D", fill_type="solid") # amarillo
        else:
            fill = PatternFill(start_color="A5D6A7", fill_type="solid") # verde

        row = [n, c_int, u, cat]
        ws.append(row)

        # Aplicar estilo a la fila recién creada
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # Ajustar ancho columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # Generar archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     as_attachment=True,
                     download_name="inventario.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
